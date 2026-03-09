"""
Exportador a Excel con formato profesional.
Genera un libro con hojas por categoría + hoja comparativa + resumen.
"""
from __future__ import annotations
from datetime import datetime
from pathlib import Path
from typing import Optional

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from models.product import ComparisonRow, Product

# ── Paleta de colores ────────────────────────────────────────────────────────
COLOR_HEADER_BG = "1F3864"   # azul oscuro Santander
COLOR_HEADER_FG = "FFFFFF"   # blanco
COLOR_BOUTIQUE = "E8F0FE"    # azul claro para filas Boutique
COLOR_GREEN = "C6EFCE"       # verde — Boutique más barato
COLOR_RED = "FFC7CE"         # rojo — Boutique más caro
COLOR_NEUTRAL = "FFFFFF"     # blanco neutro
COLOR_ALT_ROW = "F2F2F2"     # gris muy claro filas alternas
COLOR_SUBHEADER = "2E75B6"   # azul medio para sub-encabezados

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Fuentes
HEADER_FONT = Font(name="Calibri", bold=True, color=COLOR_HEADER_FG, size=11)
SUBHEADER_FONT = Font(name="Calibri", bold=True, color=COLOR_HEADER_FG, size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
BODY_FONT = Font(name="Calibri", size=10)
LINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")
DELTA_POS_FONT = Font(name="Calibri", size=10, bold=True, color="375623")  # verde oscuro
DELTA_NEG_FONT = Font(name="Calibri", size=10, bold=True, color="9C0006")  # rojo oscuro


# ── Columnas por hoja ────────────────────────────────────────────────────────

CATEGORY_COLUMNS = [
    ("Producto", 40),
    ("Marca", 12),
    ("Categoría", 14),
    ("Almacenamiento", 14),
    ("Color", 14),
    ("Fuente", 20),
    ("Tipo_Precio", 12),
    ("Precio_EUR", 12),
    ("Cuotas", 8),
    ("Precio_Mensual", 14),
    ("Disponibilidad", 14),
    ("URL", 50),
    ("Fecha_Scraping", 18),
]

COMPARISON_COLUMNS = [
    ("Producto", 42),
    ("Marca", 12),
    ("Categoría", 14),
    ("Almacenamiento", 14),
    ("Boutique_Renting_€/mes", 20),
    ("Boutique_Cuotas", 14),
    ("Rentik_€/mes", 14),
    ("Δ_Rentik_%", 12),
    ("Grover_€/mes", 14),
    ("Δ_Grover_%", 12),
    ("Movistar_€/mes", 14),
    ("Δ_Movistar_%", 12),
    ("Boutique_Compra_€", 16),
    ("Amazon_€", 12),
    ("Δ_Amazon_%", 12),
    ("MediaMarkt_€", 14),
    ("Δ_MediaMarkt_%", 14),
    ("Apple_Oficial_€", 14),
    ("Samsung_Oficial_€", 16),
    ("Boutique_URL", 45),
    ("Rentik_URL", 45),
    ("Grover_URL", 45),
    ("Movistar_URL", 45),
]

SUMMARY_COLUMNS = [
    ("Categoría", 14),
    ("Marca", 12),
    ("Nº Productos", 12),
    ("Precio Mín Boutique", 20),
    ("Precio Máx Boutique", 20),
    ("Precio Mín Mercado", 20),
    ("Precio Máx Mercado", 20),
    ("Boutique más barato (%)", 22),
    ("Boutique en precio (%)", 22),
    ("Boutique más caro (%)", 22),
]

# Mapeo de categorías a hojas
CATEGORY_SHEETS = {
    "iPhone": "iPhone",
    "iPad": "iPad & Tablets",
    "Tablet": "iPad & Tablets",
    "Mac": "Mac & Portátiles",
    "Portátil": "Mac & Portátiles",
    "Galaxy": "Samsung Galaxy",
    "AppleWatch": "Otros",
    "AirPods": "Otros",
    "TV": "Otros",
    "Otros": "Otros",
}


def export_to_excel(
    all_products: list[Product],
    comparison_rows: list[ComparisonRow],
    output_path: Optional[str] = None,
) -> str:
    """
    Genera el archivo Excel completo.
    Retorna la ruta del archivo generado.
    """
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = f"precios_santander_{timestamp}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)  # Quitar la hoja por defecto

    # 1. Hoja de resumen (primera)
    _create_summary_sheet(wb, comparison_rows, all_products)

    # 2. Hoja comparativa renting
    _create_comparison_sheet(wb, comparison_rows)

    # 3. Hojas por categoría con todos los datos brutos
    _create_category_sheets(wb, all_products)

    # 4. Hoja de metadatos
    _create_metadata_sheet(wb, all_products)

    wb.save(output_path)
    logger.success(f"Excel guardado en: {output_path}")
    return output_path


def _create_comparison_sheet(wb: Workbook, rows: list[ComparisonRow]) -> None:
    ws = wb.create_sheet("Comparativa Renting")
    _set_sheet_tab_color(ws, "2E75B6")

    # Título
    ws.merge_cells("A1:W1")
    title_cell = ws["A1"]
    title_cell.value = f"Comparativa de Precios Renting — Santander Boutique vs Competidores | {datetime.now().strftime('%d/%m/%Y')}"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Cabecera de grupos de columnas
    ws.merge_cells("A2:D2")
    _header_cell(ws, "A2", "PRODUCTO", COLOR_HEADER_BG)
    ws.merge_cells("E2:H2")
    _header_cell(ws, "E2", "RENTING MENSUAL", "1F3864")
    ws.merge_cells("I2:L2")
    _header_cell(ws, "I2", "RENTING MENSUAL", "1F3864")
    ws.merge_cells("M2:S2")
    _header_cell(ws, "M2", "PRECIO DE COMPRA", "243F60")

    # Cabecera de columnas
    row3_headers = [col for col, _ in COMPARISON_COLUMNS]
    for col_idx, header in enumerate(row3_headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = SUBHEADER_FONT
        cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_SUBHEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[3].height = 35

    # Anchos de columna
    for col_idx, (_, width) in enumerate(COMPARISON_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Datos
    delta_cols = {col: idx for idx, (col, _) in enumerate(COMPARISON_COLUMNS, 1) if col.startswith("Δ")}
    url_cols = {col: idx for idx, (col, _) in enumerate(COMPARISON_COLUMNS, 1) if col.endswith("URL")}

    for row_idx, row in enumerate(rows, 4):
        data = row.to_dict()
        is_alt = (row_idx - 4) % 2 == 1

        for col_idx, (col_name, _) in enumerate(COMPARISON_COLUMNS, 1):
            value = data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

            # Fondo alternado
            if is_alt:
                cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_ALT_ROW)

            # Destacar columnas Boutique
            if col_name in ("Boutique_Renting_€/mes", "Boutique_Compra_€"):
                cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_BOUTIQUE)
                if value:
                    cell.font = Font(name="Calibri", size=10, bold=True)
                cell.number_format = '#,##0.00 "€"'

            # Formato de precio
            elif "€" in col_name and col_name not in delta_cols and not col_name.endswith("URL"):
                if value:
                    cell.number_format = '#,##0.00 "€"'

            # Colorear deltas
            elif col_name in delta_cols:
                if isinstance(value, (int, float)) and value is not None:
                    cell.number_format = '+0.0%;-0.0%;0.0%'
                    cell.value = value / 100 if value else value
                    if value > 5:   # Boutique más barato → verde
                        cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_GREEN)
                        cell.font = DELTA_POS_FONT
                    elif value < -5:  # Boutique más caro → rojo
                        cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_RED)
                        cell.font = DELTA_NEG_FONT

            # URLs como hipervínculos
            elif col_name in url_cols and value:
                cell.value = value
                cell.font = LINK_FONT
                cell.hyperlink = value
                cell.value = "Ver →"

    # Congelar primeras 3 filas y columna A
    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(COMPARISON_COLUMNS))}3"


def _create_category_sheets(wb: Workbook, products: list[Product]) -> None:
    """Crea hojas separadas por categoría con los datos brutos."""
    # Agrupar productos por hoja
    sheets_data: dict[str, list[dict]] = {}
    for product in products:
        sheet_name = CATEGORY_SHEETS.get(product.category, "Otros")
        if sheet_name not in sheets_data:
            sheets_data[sheet_name] = []
        for row_dict in product.to_rows():
            sheets_data[sheet_name].append(row_dict)

    sheet_colors = {
        "iPhone": "1F3864",
        "Samsung Galaxy": "1428A0",
        "iPad & Tablets": "375623",
        "Mac & Portátiles": "595959",
        "Otros": "7F7F7F",
    }

    for sheet_name, rows in sheets_data.items():
        ws = wb.create_sheet(sheet_name)
        _set_sheet_tab_color(ws, sheet_colors.get(sheet_name, "1F3864"))

        if not rows:
            continue

        # Título
        n_cols = len(CATEGORY_COLUMNS)
        ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
        ws["A1"].value = f"{sheet_name} — Datos completos de precios | {datetime.now().strftime('%d/%m/%Y')}"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 24

        # Cabecera
        for col_idx, (col_name, width) in enumerate(CATEGORY_COLUMNS, 1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[2].height = 24

        # Datos
        for row_idx, row_data in enumerate(rows, 3):
            is_alt = (row_idx - 3) % 2 == 1
            source = row_data.get("Fuente", "")
            is_boutique = source == "santander_boutique"

            for col_idx, (col_name, _) in enumerate(CATEGORY_COLUMNS, 1):
                value = row_data.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")

                # Resaltar filas de Boutique
                if is_boutique:
                    cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_BOUTIQUE)
                    cell.font = Font(name="Calibri", size=10, bold=True)
                else:
                    if is_alt:
                        cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_ALT_ROW)
                    cell.font = BODY_FONT

                # Formato precio
                if col_name in ("Precio_EUR", "Precio_Mensual") and isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00 "€"'

                # URL como hipervínculo
                if col_name == "URL" and value:
                    cell.hyperlink = value
                    cell.value = "Ver →"
                    cell.font = LINK_FONT

        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}2"


def _create_summary_sheet(
    wb: Workbook, comparison_rows: list[ComparisonRow], all_products: list[Product]
) -> None:
    ws = wb.create_sheet("Resumen", 0)  # Primera hoja
    _set_sheet_tab_color(ws, "EC0000")  # Rojo Santander

    # Título principal
    ws.merge_cells("A1:J1")
    ws["A1"].value = "OBSERVATORIO DE PRECIOS — SANTANDER BOUTIQUE"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="EC0000")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:J2")
    ws["A2"].value = f"Actualizado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Total productos scrapeados: {len(all_products)}"
    ws["A2"].font = Font(name="Calibri", size=11, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # Estadísticas por fuente
    sources = list({pp.source for p in all_products for pp in p.prices})
    ws["A4"].value = "Fuentes de datos"
    ws["A4"].font = Font(name="Calibri", bold=True, size=12)
    ws.row_dimensions[4].height = 20

    ws["A5"].value = "Fuente"
    ws["B5"].value = "Nº Productos"
    for cell in [ws["A5"], ws["B5"]]:
        cell.font = HEADER_FONT
        cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center")

    source_counts: dict[str, int] = {}
    for product in all_products:
        for pp in product.prices:
            source_counts[pp.source] = source_counts.get(pp.source, 0) + 1

    for row_offset, (source, count) in enumerate(sorted(source_counts.items()), 6):
        ws.cell(row=row_offset, column=1, value=source).font = BODY_FONT
        ws.cell(row=row_offset, column=2, value=count).font = BODY_FONT

    # Tabla resumen por categoría
    start_row = 6 + len(source_counts) + 2
    ws.cell(row=start_row, column=1).value = "Resumen por Categoría"
    ws.cell(row=start_row, column=1).font = Font(name="Calibri", bold=True, size=12)

    for col_idx, (col_name, width) in enumerate(SUMMARY_COLUMNS, 1):
        cell = ws.cell(row=start_row + 1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[start_row + 1].height = 30

    # Calcular stats por categoría
    from collections import defaultdict
    cat_data: dict[tuple, list[ComparisonRow]] = defaultdict(list)
    for row in comparison_rows:
        cat_data[(row.category, row.brand)].append(row)

    for data_row, ((cat, brand), cat_rows) in enumerate(sorted(cat_data.items()), start_row + 2):
        boutique_prices = [r.boutique_renting or r.boutique_purchase for r in cat_rows if r.boutique_renting or r.boutique_purchase]
        market_prices = [
            p for r in cat_rows
            for p in [r.amazon_purchase, r.mediamarkt_purchase, r.rentik_renting, r.grover_renting]
            if p
        ]

        cheaper = sum(1 for r in cat_rows if _boutique_is_cheaper(r))
        same = sum(1 for r in cat_rows if _boutique_is_same(r))
        pricier = sum(1 for r in cat_rows if _boutique_is_pricier(r))
        total = len(cat_rows)

        values = [
            cat, brand, total,
            min(boutique_prices) if boutique_prices else None,
            max(boutique_prices) if boutique_prices else None,
            min(market_prices) if market_prices else None,
            max(market_prices) if market_prices else None,
            f"{cheaper/total*100:.0f}%" if total else "N/A",
            f"{same/total*100:.0f}%" if total else "N/A",
            f"{pricier/total*100:.0f}%" if total else "N/A",
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=data_row, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if isinstance(value, float) and col_idx in (4, 5, 6, 7):
                cell.number_format = '#,##0.00 "€"'


def _create_metadata_sheet(wb: Workbook, all_products: list[Product]) -> None:
    ws = wb.create_sheet("Metadatos")
    ws["A1"].value = "Información del scraping"
    ws["A1"].font = TITLE_FONT
    ws["A2"].value = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A3"].value = f"Total de productos: {len(all_products)}"
    ws["A4"].value = f"Total de registros de precio: {sum(len(p.prices) for p in all_products)}"
    ws["A5"].value = "Sistema: Santander Boutique Price Comparator v1.0"
    ws["A6"].value = "Motor de scraping: Scrapling (Fetcher + StealthyFetcher + DynamicFetcher)"
    for row in range(1, 7):
        ws.cell(row=row, column=1).font = BODY_FONT
    ws.column_dimensions["A"].width = 60


# ── Helpers ──────────────────────────────────────────────────────────────────

def _header_cell(ws, cell_ref: str, value: str, color: str) -> None:
    cell = ws[cell_ref]
    cell.value = value
    cell.font = HEADER_FONT
    cell.fill = PatternFill(fill_type="solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER


def _set_sheet_tab_color(ws, color: str) -> None:
    ws.sheet_properties.tabColor = color


def _boutique_is_cheaper(row: ComparisonRow) -> bool:
    boutique = row.boutique_renting or row.boutique_purchase
    if not boutique:
        return False
    competitors = [
        p for p in [row.rentik_renting, row.grover_renting, row.amazon_purchase, row.mediamarkt_purchase]
        if p
    ]
    if not competitors:
        return False
    return boutique < min(competitors)


def _boutique_is_same(row: ComparisonRow) -> bool:
    boutique = row.boutique_renting or row.boutique_purchase
    if not boutique:
        return False
    competitors = [
        p for p in [row.rentik_renting, row.grover_renting, row.amazon_purchase, row.mediamarkt_purchase]
        if p
    ]
    if not competitors:
        return False
    return abs(boutique - min(competitors)) / boutique < 0.05


def _boutique_is_pricier(row: ComparisonRow) -> bool:
    boutique = row.boutique_renting or row.boutique_purchase
    if not boutique:
        return False
    competitors = [
        p for p in [row.rentik_renting, row.grover_renting, row.amazon_purchase, row.mediamarkt_purchase]
        if p
    ]
    if not competitors:
        return False
    return boutique > min(competitors) * 1.05
